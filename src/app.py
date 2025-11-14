import re
from io import BytesIO

import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font


# ============= PARSER EUROFIEL =============

def split_orders(full_text: str):
    """
    Divide el texto completo del PDF en bloques,
    cada uno correspondiente a un pedido (PEDIDO / REEMPLAZO / ANULACIÓN).
    """
    matches = list(re.finditer(r"Nº Pedido\s*:", full_text))
    chunks = []

    for i, m in enumerate(matches):
        start = m.start()
        # Buscamos la línea anterior para incluir el TIPO (PEDIDO, ANULACIÓN PEDIDO…)
        prev_nl = full_text.rfind("\n", 0, start)
        if prev_nl != -1:
            prev_prev_nl = full_text.rfind("\n", 0, prev_nl)
            order_start = prev_prev_nl + 1 if prev_prev_nl != -1 else 0
        else:
            order_start = 0

        end = matches[i + 1].start() if i + 1 < len(matches) else len(full_text)
        chunk = full_text[order_start:end].strip()
        chunks.append(chunk)

    return chunks


def parse_detail_line_eurofiel(line: str):
    """
    Parsea una línea de detalle de artículo Eurofiel.

    Ej:
      1 8447571299747 3RC240/NARANJA/XS 0863769/66/01 1 50 50 0 EUR
    Devuelve (MODELO, PATRON, PRECIO) o None si la línea no es de detalle.
    """
    parts = line.split()
    if len(parts) < 8:
        return None

    if not parts[0].isdigit():
        return None

    if not re.fullmatch(r"\d{13}", parts[1]):
        return None

    cli_idx = None
    for i in range(2, len(parts)):
        if re.fullmatch(r"\d+/\d+/\d+", parts[i]):
            cli_idx = i
            break

    if cli_idx is None or cli_idx + 4 >= len(parts):
        return None

    cod_prov_full = " ".join(parts[2:cli_idx])
    cod_cli_full = parts[cli_idx]

    p_neto = parts[cli_idx + 3]

    # MODELO = Cod Proveedor/Color (quitamos talla)
    modelo = re.sub(r"/[^/]+$", "", cod_prov_full)
    # PATRON = Cod Cliente/Color (quitamos talla)
    patron = re.sub(r"/[^/]+$", "", cod_cli_full)

    precio = p_neto.replace(",", ".")

    return modelo, patron, precio


def parse_order_eurofiel(order_text: str):
    """
    Parsea un bloque de texto correspondiente a un solo pedido Eurofiel.
    """
    lines = [ln for ln in order_text.splitlines() if ln.strip()]
    first_line = lines[0].strip() if lines else ""
    tipo = first_line  # PEDIDO / REEMPLAZO PEDIDO / ANULACIÓN PEDIDO

    def search(pattern: str):
        m = re.search(pattern, order_text)
        return m.group(1).strip() if m else ""

    pedido = search(r"Nº Pedido\s*:\s*(\S+)")
    fecha_entrega = search(r"Fecha Entrega\s*:\s*(\d{2}/\d{2}/\d{4})")

    pais = ""
    m_pais = re.search(r"País:\s*\([^)]*\)\s*([A-ZÁÉÍÓÚÜÑ ]+)", order_text)
    if m_pais:
        pais = m_pais.group(1).strip()

    descripcion = search(r"Descripción:\s*(.+)")
    total_unidades = search(r"Total Unidades\s+(\d+)")

    modelo = ""
    patron = ""
    precio = ""

    for ln in lines:
        parsed = parse_detail_line_eurofiel(ln)
        if parsed:
            modelo, patron, precio = parsed
            break

    return {
        "TIPO": tipo,
        "PEDIDO": pedido,
        "FECHA_ENTREGA": fecha_entrega,
        "PAIS": pais,
        "DESCRIPCION": descripcion,
        "MODELO": modelo,
        "PATRON": patron,
        "PRECIO": precio,
        "TOTAL_UNIDADES": total_unidades,
    }


def parse_pdf_eurofiel_bytes(pdf_bytes: bytes):
    """
    Parsea un PDF Eurofiel que viene en memoria (subido por web).
    """
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    orders = split_orders(full_text)
    rows = [parse_order_eurofiel(o) for o in orders]
    rows = [r for r in rows if r.get("PEDIDO")]  # limpia ruido
    return pd.DataFrame(rows)


# ============= PARSER ECI =============

def parse_page_eci(text: str):
    """
    Parsea una página de pedido de ECI.
    Devuelve una lista de dicts, una fila por (pedido, modelo, color).
    """

    # Limpiamos líneas
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def search(pattern: str):
        m = re.search(pattern, text)
        return m.group(1).strip() if m else ""

    # TIPO: Pedido / Reposición / Anulación Pedido...
    tipo = ""
    for ln in lines:
        low = ln.lower()
        if low in ("pedido", "reposicion", "reposición", "anulacion pedido", "anulación pedido"):
            tipo = ln.upper()
            break

    # Cabecera
    n_pedido = search(r"Nº Pedido\s+(\d+)")
    departamento = search(r"Dpto\. venta\s+(\d+)")
    fecha_entrega = search(r"Fecha Entrega\s+(\d{2}/\d{2}/\d{4})")

    # Sucursal entrega (01 0050, 02 0062, etc.)
    suc_entrega = search(r"Sucursal Destino que Pide\s+([0-9 ]+)\s+[A-ZÁÉÍÓÚÜÑ]")
    if not suc_entrega:
        suc_entrega = search(r"Sucursal de Entrega\s+([0-9 ]+)\s+[A-ZÁÉÍÓÚÜÑ]")

    rows = []

    for i, ln in enumerate(lines):
        # Línea de detalle principal (nº + EAN13 + resto)
        if not re.match(r"^\d+\s+\d{13}\s", ln):
            continue

        parts = ln.split()

        # Índices de tokens numéricos (nos sirven para localizar QTY y precios)
        num_indices = [idx for idx, tok in enumerate(parts)
                       if re.fullmatch(r"[\d.,]+", tok)]
        if len(num_indices) < 6:
            continue

        # Patrón: ... DESCRIPCION QTY 1 P_BRUTO P_NETO PVP NETO_LINEA
        qty_idx = num_indices[-6]
        p_bruto_idx = num_indices[-4]

        qty = parts[qty_idx]
        p_bruto_raw = parts[p_bruto_idx]

        # Normalizamos precio: 53,000 → 53.000
        precio = p_bruto_raw.replace(".", "").replace(",", ".")

        # Descripción: después de los 3 códigos (serie + ref + colorcode)
        desc_tokens = parts[5:qty_idx]
        descripcion = " ".join(desc_tokens)

        # Línea siguiente puede ser segunda parte de la descripción (PUNTO ASIM FALDA)
        extra_desc = ""
        if i + 1 < len(lines):
            next_ln = lines[i + 1]
            if (not re.match(r"^\d+\s+\d{13}\s", next_ln)
                and not re.match(r"^[A-Z0-9]{5,}\s+\d{3}\s", next_ln)
                and "WOMAN FIESTA" not in next_ln):
                extra_desc = next_ln.strip()

        if extra_desc:
            descripcion = f"{descripcion} {extra_desc}"

        # Línea de modelo/color:
        # si hay extra_desc, está 2 líneas debajo; si no, 1 línea debajo
        j = i + 1 + (1 if extra_desc else 0)
        modelo = ""
        color = ""

        if j < len(lines):
            info_ln = lines[j]
            info_parts = info_ln.split()

            # Formato visto: 47D262G 983 PRINT NEGRO003 3
            if len(info_parts) >= 3 and re.fullmatch(r"[A-Z0-9]+", info_parts[0]):
                modelo = info_parts[0]
                color_code = info_parts[1] if len(info_parts) >= 2 else ""
                color_name1 = info_parts[2] if len(info_parts) >= 3 else ""
                color_name2 = ""
                if len(info_parts) >= 4:
                    color_name2 = re.sub(r"\d+$", "", info_parts[3])
                color = " ".join([p for p in [color_code, color_name1, color_name2] if p])

        rows.append({
            "TIPO": tipo,
            "N_PEDIDO": n_pedido,
            "DEPARTAMENTO": departamento,
            "DESCRIPCION": descripcion,
            "MODELO": modelo,
            "COLOR": color,
            "PRECIO": precio,
            "FECHA_ENTREGA": fecha_entrega,
            "SUC_ENTREGA": suc_entrega,
            "TOTAL_UNIDADES": qty,
        })

    return rows


def parse_pdf_eci_bytes(pdf_bytes: bytes) -> pd.DataFrame:
    """
    Abre un PDF de ECI (bytes) y devuelve un DataFrame con:
    - 1 fila por (PEDIDO + MODELO + COLOR)
    - TOTAL_UNIDADES: suma por modelo/color/pedido
    """
    all_rows = []

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_rows.extend(parse_page_eci(text))

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    df["TOTAL_UNIDADES"] = pd.to_numeric(
        df["TOTAL_UNIDADES"], errors="coerce"
    ).fillna(0).astype(int)

    group_cols = [
        "TIPO",
        "N_PEDIDO",
        "DEPARTAMENTO",
        "DESCRIPCION",
        "MODELO",
        "COLOR",
        "PRECIO",
        "FECHA_ENTREGA",
        "SUC_ENTREGA",
    ]

    df_grouped = (
        df.groupby(group_cols, as_index=False)["TOTAL_UNIDADES"]
          .sum()
    )

    return df_grouped


# ============= UTILIDADES COMUNES (COLORES + BORDES + TOTALES) =============

PALETTE = [
    "#fde2e4",
    "#bee1e6",
    "#e2f0cb",
    "#ffdfba",
    "#d0bdf4",
    "#c7f9cc",
    "#ffccd5",
    "#f1f0ff",
    "#e5f4e3",
    "#ffe5b4",
    "#e0bbff",
    "#caffbf",
    "#ffd6a5",
    "#bde0fe",
    "#ffafcc",
]


def style_by_model(df: pd.DataFrame):
    modelos = df["MODELO"].fillna("").astype(str).unique()
    model_colors = {}
    for i, m in enumerate(modelos):
        if not m:
            continue
        color = PALETTE[i % len(PALETTE)]
        model_colors[m] = color

    def color_rows(row):
        color = model_colors.get(str(row["MODELO"]), "")
        if not color:
            return [""] * len(row)
        return [f"background-color: {color}"] * len(row)

    return df.style.apply(color_rows, axis=1)


def style_workbook_with_borders_and_headers(workbook):
    """
    Aplica:
    - Bordes finos a todas las celdas
    - Cabeceras en amarillo chillón + negrita (fila 1)
    - Filas cuyo primer valor sea 'TOTAL' → amarillo + negrita
    """
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="FFFF00")  # amarillo fuerte
    header_font = Font(bold=True)

    total_fill = PatternFill("solid", fgColor="FFFF00")
    total_font = Font(bold=True)

    for ws in workbook.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # Bordes + cabecera
        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # Totales (fila con primera celda == 'TOTAL')
        for row_idx in range(2, max_row + 1):
            first_cell = ws.cell(row=row_idx, column=1)
            if str(first_cell.value).strip().upper() == "TOTAL":
                for col_idx in range(1, max_col + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.fill = total_fill
                    c.font = total_font


# ============= STREAMLIT APP =============

st.set_page_config(page_title="Resumen pedidos EDIWIN", layout="wide")

st.title("📦 Resumen de pedidos desde PDF EDIWIN")

cliente = st.selectbox("Cliente", ["Eurofiel", "El Corte Inglés"])

st.write(
    "Sube un PDF de **Eurofiel** o **El Corte Inglés** y te saco "
    "las líneas clave listas para Excel."
)

label = "📁 Sube tu PDF de Eurofiel" if cliente == "Eurofiel" else "📁 Sube tu PDF de El Corte Inglés"
uploaded_pdf = st.file_uploader(label, type=["pdf"])

if uploaded_pdf is not None:
    try:
        if cliente == "Eurofiel":
            df = parse_pdf_eurofiel_bytes(uploaded_pdf.getvalue())
        else:
            df = parse_pdf_eci_bytes(uploaded_pdf.getvalue())

        if df.empty:
            st.warning("No se han detectado pedidos en el PDF. Revisa el formato.")
        else:
            st.subheader("📊 Vista previa de pedidos detectados")

            # Aseguramos TOTAL_UNIDADES numérico
            if "TOTAL_UNIDADES" in df.columns:
                df["TOTAL_UNIDADES"] = pd.to_numeric(
                    df["TOTAL_UNIDADES"], errors="coerce"
                ).fillna(0)

            # Estilos por MODELO
            styled_df = style_by_model(df)
            st.dataframe(styled_df, use_container_width=True)

            # ====== RESÚMENES Y EXPORT ======
            if cliente == "Eurofiel":
                st.subheader("📦 Resumen por MODELO")

                resumen = (
                    df.groupby("MODELO", dropna=False)
                    .agg(
                        PEDIDOS=("PEDIDO", "nunique"),
                        UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
                    )
                    .reset_index()
                    .sort_values("PEDIDOS", ascending=False)
                )

                st.dataframe(resumen, use_container_width=True)

                # --- versión para Excel con fila TOTAL ---
                resumen_xlsx = resumen.copy()
                total_row = {
                    "MODELO": "TOTAL",
                    "PEDIDOS": resumen_xlsx["PEDIDOS"].sum(),
                    "UNIDADES_TOTALES": resumen_xlsx["UNIDADES_TOTALES"].sum(),
                }
                resumen_xlsx = pd.concat(
                    [resumen_xlsx, pd.DataFrame([total_row])],
                    ignore_index=True,
                )

                # ---- Exportar Excel Eurofiel ----
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                    styled_df.to_excel(writer, sheet_name="Pedidos", index=False)
                    resumen_xlsx.to_excel(writer, sheet_name="Resumen por modelo", index=False)

                    wb = writer.book
                    style_workbook_with_borders_and_headers(wb)

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_buffer.getvalue(),
                    file_name="eurofiel_resumen_pedidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # CSV simple
                csv_bytes = df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Descargar CSV",
                    data=csv_bytes,
                    file_name="eurofiel_resumen_pedidos.csv",
                    mime="text/csv",
                )

            else:  # El Corte Inglés
                st.subheader("📦 Resumen por MODELO + COLOR")

                resumen_mc = (
                    df.groupby(["MODELO", "COLOR"], dropna=False)
                    .agg(
                        PEDIDOS=("N_PEDIDO", "nunique"),
                        UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
                    )
                    .reset_index()
                    .sort_values("PEDIDOS", ascending=False)
                )

                st.dataframe(resumen_mc, use_container_width=True)

                st.subheader("🧩 Resumen por MODELO (todas las sucursales/colores)")

                resumen_m = (
                    df.groupby("MODELO", dropna=False)
                    .agg(
                        PEDIDOS=("N_PEDIDO", "nunique"),
                        UNIDADES_TOTALES=("TOTAL_UNIDADES", "sum"),
                    )
                    .reset_index()
                    .sort_values("PEDIDOS", ascending=False)
                )

                st.dataframe(resumen_m, use_container_width=True)

                # --- versiones para Excel con fila TOTAL ---
                resumen_mc_xlsx = resumen_mc.copy()
                total_mc = {
                    "MODELO": "TOTAL",
                    "COLOR": "",
                    "PEDIDOS": resumen_mc_xlsx["PEDIDOS"].sum(),
                    "UNIDADES_TOTALES": resumen_mc_xlsx["UNIDADES_TOTALES"].sum(),
                }
                resumen_mc_xlsx = pd.concat(
                    [resumen_mc_xlsx, pd.DataFrame([total_mc])],
                    ignore_index=True,
                )

                resumen_m_xlsx = resumen_m.copy()
                total_m = {
                    "MODELO": "TOTAL",
                    "PEDIDOS": resumen_m_xlsx["PEDIDOS"].sum(),
                    "UNIDADES_TOTALES": resumen_m_xlsx["UNIDADES_TOTALES"].sum(),
                }
                resumen_m_xlsx = pd.concat(
                    [resumen_m_xlsx, pd.DataFrame([total_m])],
                    ignore_index=True,
                )

                # ---- Exportar Excel ECI ----
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                    styled_df.to_excel(writer, sheet_name="Pedidos", index=False)
                    resumen_mc_xlsx.to_excel(writer, sheet_name="Resumen modelo+color", index=False)
                    resumen_m_xlsx.to_excel(writer, sheet_name="Resumen modelo", index=False)

                    wb = writer.book
                    style_workbook_with_borders_and_headers(wb)

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_buffer.getvalue(),
                    file_name="eci_resumen_pedidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # CSV simple
                csv_bytes = df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Descargar CSV",
                    data=csv_bytes,
                    file_name="eci_resumen_pedidos.csv",
                    mime="text/csv",
                )

    except Exception as e:
        st.error(f"❌ Error procesando el PDF: {e}")
else:
    st.info("Sube un PDF para empezar.")


# ===== FOOTER =====
st.markdown("""
<hr style="margin-top: 60px;">

<div style="text-align:center; font-size:14px; color:#777;">
    Creado y desarrollado con mucho amor ❤️ por<br>
    <strong style="font-size:16px;">Aitor Susperregui</strong><br>
    <span style="font-size:12px;">@elvasco.x</span>
</div>
""", unsafe_allow_html=True)
